"""
FastAPI 主應用
NOTE: 提供 member 資料表的 CRUD API
"""
import logging
import os
import io
from typing import List, Optional
from enum import Enum

from fastapi import FastAPI, HTTPException, Query, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from starlette.responses import JSONResponse

from database import get_cursor
from models import (
    Member, MemberCreate, MemberUpdate,
    Attendance, AttendanceCreate, AttendanceUpdate,
    CodeTable, CodeTableCreate, CodeTableUpdate,
    AnnualLeave, AnnualLeaveCreate, AnnualLeaveUpdate,
    ProjectInfo, ProjectInfoCreate, ProjectInfoUpdate
)
from csrf import (
    CSRFMiddleware,
    generate_csrf_token,
    set_csrf_cookie,
    CSRF_COOKIE_NAME,
)
from routes import auth_router, system_router

# 設定日誌
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(
    title="部門人員管理 API",
    description="提供員工資料的 CRUD 操作",
    version="1.0.0",
)

# CORS 設定 - 允許前端跨域請求
# NOTE: 在 Docker 環境中，前端透過 Nginx 反向代理存取 API
allowed_origins = [
    "http://localhost:5173",
    "http://127.0.0.1:5173",
    "http://localhost:5174",
    "http://127.0.0.1:5174",
    "http://localhost",
    "http://127.0.0.1",
    "http://localhost:80",
    "http://127.0.0.1:80",
]

# 從環境變數讀取額外的允許來源
extra_origins = os.getenv("CORS_ORIGINS", "")
if extra_origins:
    allowed_origins.extend(extra_origins.split(","))

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["X-CSRF-Token"],  # 允許前端讀取 CSRF header
)

# 添加 CSRF 中間件
# NOTE: CSRF 中間件必須在 CORS 之後添加
app.add_middleware(CSRFMiddleware)

# 註冊認證與系統管理路由
app.include_router(auth_router)
app.include_router(system_router)

# 註冊公告路由
from routes import announcement_router
app.include_router(announcement_router)

@app.get("/health")
def health_check():
    """
    健康檢查端點
    NOTE: 用於 Docker 健康檢查和負載平衡器探測
    """
    return {"status": "healthy", "service": "section-builder-api"}


@app.get("/")
def root():
    """
    API 根路徑
    """
    return {"message": "部門人員管理 API", "version": "1.0.0"}


@app.get("/api/csrf-token")
def get_csrf_token():
    """
    取得 CSRF Token
    NOTE: 前端在應用啟動時調用此端點獲取 token，
          token 會同時設置在 Cookie 和回應 body 中
    """
    token = generate_csrf_token()
    response = JSONResponse(content={"csrf_token": token})
    set_csrf_cookie(response, token)
    return response


@app.get("/api/members")
def get_members(
    search: Optional[str] = Query(None, description="搜尋關鍵字"),
    division: Optional[str] = Query(None, description="部門篩選"),
    is_employed: Optional[bool] = Query(None, description="在職狀態篩選"),
    member_type: Optional[List[str]] = Query(None, description="員工類型篩選，可多選 (member/manager/intern/consultant/outsourcing)"),
    page: int = Query(1, ge=1, description="頁碼"),
    page_size: int = Query(20, ge=1, le=100, description="每頁筆數"),
    sort_by: Optional[str] = Query(None, description="排序欄位"),
    sort_order: Optional[str] = Query("asc", description="排序方向 asc/desc"),
):
    """
    取得所有員工清單
    支援分頁、排序和篩選
    """
    try:
        with get_cursor() as cursor:
            # 建立 SQL 查詢
            base_sql = "SELECT * FROM member WHERE 1=1"
            params = []

            if search:
                base_sql += """ AND (
                    chinese_name ILIKE %s OR 
                    name ILIKE %s OR 
                    emp_id ILIKE %s OR
                    email ILIKE %s OR
                    job_title ILIKE %s
                )"""
                search_pattern = f"%{search}%"
                params.extend([search_pattern] * 5)

            if division:
                base_sql += " AND division_name = %s"
                params.append(division)

            if is_employed is not None:
                base_sql += " AND is_employed = %s"
                params.append(is_employed)

            # 員工類型篩選（支援多選，使用 OR 條件）
            if member_type and len(member_type) > 0:
                type_column_map = {
                    'member': 'is_member',
                    'manager': 'is_manager',
                    'intern': 'is_intern',
                    'consultant': 'is_consultant',
                    'outsourcing': 'is_outsourcing',
                }
                # 過濾有效的類型
                valid_types = [t for t in member_type if t in type_column_map]
                if valid_types:
                    type_conditions = [f"{type_column_map[t]} = true" for t in valid_types]
                    base_sql += f" AND ({' OR '.join(type_conditions)})"

            # 計算總筆數
            count_sql = f"SELECT COUNT(*) as total FROM ({base_sql}) as subquery"
            cursor.execute(count_sql, params)
            total = cursor.fetchone()['total']

            # 排序
            allowed_sort_fields = ['emp_id', 'chinese_name', 'name', 'division_name', 'job_title', 'email']
            if sort_by and sort_by in allowed_sort_fields:
                order_direction = 'ASC' if sort_order == 'asc' else 'DESC'
                base_sql += f" ORDER BY {sort_by} {order_direction}"
            else:
                base_sql += " ORDER BY emp_id"

            # 分頁
            offset = (page - 1) * page_size
            base_sql += " LIMIT %s OFFSET %s"
            params.extend([page_size, offset])

            cursor.execute(base_sql, params)
            rows = cursor.fetchall()

            total_pages = (total + page_size - 1) // page_size

            return {
                "items": [dict(row) for row in rows],
                "total": total,
                "page": page,
                "page_size": page_size,
                "total_pages": total_pages,
            }

    except Exception as e:
        logger.error(f"取得員工清單失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/members/stats")
def get_member_stats(
    dimension: str = Query("member_type", description="統計維度: member_type, division, job_title"),
    member_types: Optional[str] = Query(None, description="員工類型（逗號分隔）: member,manager,intern,consultant,outsourcing"),
    employed_status: Optional[str] = Query(None, description="在職狀態（逗號分隔）: employed,unemployed"),
):
    """
    取得員工統計資料
    NOTE: 支援依員工類型、部門、職稱統計，可多選篩選
    """
    try:
        with get_cursor() as cursor:
            # 員工類型欄位對照
            type_column_map = {
                'member': 'is_member',
                'manager': 'is_manager',
                'intern': 'is_intern',
                'consultant': 'is_consultant',
                'outsourcing': 'is_outsourcing'
            }
            
            # 建立基礎條件
            conditions = []
            
            # 員工類型篩選
            selected_types = []
            if member_types:
                selected_types = [t.strip() for t in member_types.split(",") if t.strip() in type_column_map]
            
            if selected_types:
                type_conditions = [f"{type_column_map[t]} = TRUE" for t in selected_types]
                conditions.append(f"({' OR '.join(type_conditions)})")
            
            # 在職狀態篩選
            if employed_status:
                status_list = [s.strip() for s in employed_status.split(",") if s.strip()]
                if 'employed' in status_list and 'unemployed' not in status_list:
                    conditions.append("is_employed = TRUE")
                elif 'unemployed' in status_list and 'employed' not in status_list:
                    conditions.append("is_employed = FALSE")
                # 如果兩者都選，則不加條件（顯示全部）
            
            where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""
            
            if dimension == "member_type":
                # 員工類型需要特殊處理（多個布林欄位）
                items = []
                types_to_count = selected_types if selected_types else list(type_column_map.keys())
                
                for label in types_to_count:
                    col = type_column_map[label]
                    # 構建條件：員工類型 + 在職狀態
                    type_conditions = [f"{col} = TRUE"]
                    
                    if employed_status:
                        status_list = [s.strip() for s in employed_status.split(",") if s.strip()]
                        if 'employed' in status_list and 'unemployed' not in status_list:
                            type_conditions.append("is_employed = TRUE")
                        elif 'unemployed' in status_list and 'employed' not in status_list:
                            type_conditions.append("is_employed = FALSE")
                    
                    type_where = f"WHERE {' AND '.join(type_conditions)}"
                    
                    query = f"""
                        SELECT COUNT(*) as count
                        FROM member
                        {type_where}
                    """
                    cursor.execute(query)
                    row = cursor.fetchone()
                    count = row["count"] if row else 0
                    if count > 0:
                        items.append({"name": label, "count": count})
                
                # 按數量排序
                items.sort(key=lambda x: x["count"], reverse=True)
                return {"items": items}
            
            elif dimension == "division":
                group_field = "COALESCE(division_name, '未設定')"
            elif dimension == "job_title":
                group_field = "COALESCE(job_title, '未設定')"
            else:
                group_field = "COALESCE(division_name, '未設定')"
            
            query = f"""
                SELECT 
                    {group_field} as name,
                    COUNT(*) as count
                FROM member
                {where_clause}
                GROUP BY {group_field}
                ORDER BY count DESC
            """
            
            cursor.execute(query)
            rows = cursor.fetchall()
            
            items = [
                {
                    "name": row["name"] if row["name"] else "未設定",
                    "count": row["count"]
                }
                for row in rows
            ]
            
            return {"items": items}
            
    except Exception as e:
        logger.error(f"統計員工失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/members/{emp_id}", response_model=Member)
def get_member(emp_id: str):
    """
    根據員工編號取得單一員工資料
    """
    try:
        with get_cursor() as cursor:
            cursor.execute("SELECT * FROM member WHERE emp_id = %s", (emp_id,))
            row = cursor.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="員工不存在")
            return dict(row)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"取得員工資料失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/members", response_model=Member)
def create_member(member: MemberCreate):
    """
    新增員工
    """
    try:
        with get_cursor() as cursor:
            # 檢查員工編號是否已存在
            cursor.execute("SELECT emp_id FROM member WHERE emp_id = %s", (member.emp_id,))
            if cursor.fetchone():
                raise HTTPException(status_code=400, detail="員工編號已存在")

            # 取得非 None 的欄位
            data = member.model_dump(exclude_none=True)
            columns = ", ".join(data.keys())
            placeholders = ", ".join(["%s"] * len(data))
            values = list(data.values())

            sql = f"INSERT INTO member ({columns}) VALUES ({placeholders}) RETURNING *"
            cursor.execute(sql, values)
            row = cursor.fetchone()
            return dict(row)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"新增員工失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/members/{emp_id}", response_model=Member)
def update_member(emp_id: str, member: MemberUpdate):
    """
    更新員工資料
    """
    try:
        with get_cursor() as cursor:
            # 檢查員工是否存在
            cursor.execute("SELECT emp_id FROM member WHERE emp_id = %s", (emp_id,))
            if not cursor.fetchone():
                raise HTTPException(status_code=404, detail="員工不存在")

            # 取得非 None 的欄位進行更新
            data = member.model_dump(exclude_none=True)
            if not data:
                raise HTTPException(status_code=400, detail="沒有要更新的欄位")

            set_clause = ", ".join([f"{k} = %s" for k in data.keys()])
            values = list(data.values()) + [emp_id]

            sql = f"UPDATE member SET {set_clause} WHERE emp_id = %s RETURNING *"
            cursor.execute(sql, values)
            row = cursor.fetchone()
            return dict(row)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新員工失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/members/{emp_id}")
def delete_member(emp_id: str):
    """
    刪除員工
    """
    try:
        with get_cursor() as cursor:
            cursor.execute("DELETE FROM member WHERE emp_id = %s RETURNING emp_id", (emp_id,))
            row = cursor.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="員工不存在")
            return {"message": "刪除成功", "emp_id": emp_id}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"刪除員工失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/divisions", response_model=List[str])
def get_divisions():
    """
    取得所有部門清單（用於篩選下拉選單）
    """
    try:
        with get_cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT division_name 
                FROM member 
                WHERE division_name IS NOT NULL AND division_name != ''
                ORDER BY division_name
            """)
            rows = cursor.fetchall()
            return [row["division_name"] for row in rows]

    except Exception as e:
        logger.error(f"取得部門清單失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# Attendance (請假記錄) API 端點
# ============================================

@app.get("/api/attendance")
def get_attendance_records(
    emp_id: Optional[str] = Query(None, description="員工編號篩選"),
    emp_name: Optional[str] = Query(None, description="員工姓名篩選"),
    leave_type: Optional[str] = Query(None, description="假別篩選"),
    start_date: Optional[str] = Query(None, description="開始日期 YYYYMMDD"),
    end_date: Optional[str] = Query(None, description="結束日期 YYYYMMDD"),
    page: int = Query(1, ge=1, description="頁碼"),
    page_size: int = Query(20, ge=1, le=100, description="每頁筆數"),
    sort_by: Optional[str] = Query(None, description="排序欄位"),
    sort_order: Optional[str] = Query("desc", description="排序方向 asc/desc"),
):
    """
    取得請假記錄清單
    支援分頁、排序和篩選
    """
    try:
        with get_cursor() as cursor:
            # 基礎查詢（JOIN 員工資料取得姓名）
            base_sql = """
                SELECT a.*, m.chinese_name, m.name as english_name
                FROM member_attendance a
                LEFT JOIN member m ON a.emp_id = m.emp_id
                WHERE 1=1
            """
            params = []

            if emp_id:
                base_sql += " AND a.emp_id = %s"
                params.append(emp_id)

            if emp_name:
                base_sql += " AND (m.chinese_name LIKE %s OR m.name LIKE %s)"
                params.append(f"%{emp_name}%")
                params.append(f"%{emp_name}%")

            if leave_type:
                base_sql += " AND a.leave_type = %s"
                params.append(leave_type)

            if start_date:
                base_sql += " AND a.leave_date >= %s"
                params.append(start_date)

            if end_date:
                base_sql += " AND a.leave_date <= %s"
                params.append(end_date)

            # 計算總筆數
            count_sql = f"SELECT COUNT(*) as total FROM ({base_sql}) as subquery"
            cursor.execute(count_sql, params)
            total = cursor.fetchone()['total']

            # 排序
            allowed_sort_fields = ['emp_id', 'leave_date', 'leave_type', 'duration_days', 'chinese_name']
            if sort_by and sort_by in allowed_sort_fields:
                order_direction = 'ASC' if sort_order == 'asc' else 'DESC'
                base_sql += f" ORDER BY {sort_by} {order_direction}"
            else:
                base_sql += " ORDER BY a.leave_date DESC, a.emp_id"

            # 分頁
            offset = (page - 1) * page_size
            base_sql += " LIMIT %s OFFSET %s"
            params.extend([page_size, offset])

            cursor.execute(base_sql, params)
            rows = cursor.fetchall()

            total_pages = (total + page_size - 1) // page_size

            return {
                "items": [dict(row) for row in rows],
                "total": total,
                "page": page,
                "page_size": page_size,
                "total_pages": total_pages,
            }

    except Exception as e:
        logger.error(f"取得請假記錄失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))




@app.get("/api/attendance/{emp_id}/{leave_date}/{leave_type}", response_model=Attendance)
def get_attendance_record(emp_id: str, leave_date: str, leave_type: str):
    """
    根據複合主鍵取得單一請假記錄
    """
    try:
        with get_cursor() as cursor:
            cursor.execute(
                "SELECT * FROM member_attendance WHERE emp_id = %s AND leave_date = %s AND leave_type = %s",
                (emp_id, leave_date, leave_type)
            )
            row = cursor.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="請假記錄不存在")
            return dict(row)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"取得請假記錄失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/attendance", response_model=Attendance)
def create_attendance_record(attendance: AttendanceCreate):
    """
    新增請假記錄
    """
    try:
        with get_cursor() as cursor:
            # 檢查記錄是否已存在
            cursor.execute(
                "SELECT emp_id FROM member_attendance WHERE emp_id = %s AND leave_date = %s AND leave_type = %s",
                (attendance.emp_id, attendance.leave_date, attendance.leave_type)
            )
            if cursor.fetchone():
                raise HTTPException(status_code=400, detail="此請假記錄已存在")

            # 取得非 None 的欄位
            data = attendance.model_dump(exclude_none=True)
            columns = ", ".join(data.keys())
            placeholders = ", ".join(["%s"] * len(data))
            values = list(data.values())

            sql = f"INSERT INTO member_attendance ({columns}) VALUES ({placeholders}) RETURNING *"
            cursor.execute(sql, values)
            row = cursor.fetchone()
            return dict(row)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"新增請假記錄失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/attendance/{emp_id}/{leave_date}/{leave_type}", response_model=Attendance)
def update_attendance_record(emp_id: str, leave_date: str, leave_type: str, attendance: AttendanceUpdate):
    """
    更新請假記錄
    """
    try:
        with get_cursor() as cursor:
            # 檢查記錄是否存在
            cursor.execute(
                "SELECT emp_id FROM member_attendance WHERE emp_id = %s AND leave_date = %s AND leave_type = %s",
                (emp_id, leave_date, leave_type)
            )
            if not cursor.fetchone():
                raise HTTPException(status_code=404, detail="請假記錄不存在")

            # 取得非 None 的欄位進行更新
            data = attendance.model_dump(exclude_none=True)
            if not data:
                raise HTTPException(status_code=400, detail="沒有要更新的欄位")

            set_clause = ", ".join([f"{k} = %s" for k in data.keys()])
            values = list(data.values()) + [emp_id, leave_date, leave_type]

            sql = f"UPDATE member_attendance SET {set_clause} WHERE emp_id = %s AND leave_date = %s AND leave_type = %s RETURNING *"
            cursor.execute(sql, values)
            row = cursor.fetchone()
            return dict(row)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新請假記錄失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/attendance/{emp_id}/{leave_date}/{leave_type}")
def delete_attendance_record(emp_id: str, leave_date: str, leave_type: str):
    """
    刪除請假記錄
    """
    try:
        with get_cursor() as cursor:
            cursor.execute(
                "DELETE FROM member_attendance WHERE emp_id = %s AND leave_date = %s AND leave_type = %s RETURNING emp_id",
                (emp_id, leave_date, leave_type)
            )
            row = cursor.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="請假記錄不存在")
            return {"message": "刪除成功", "emp_id": emp_id, "leave_date": leave_date, "leave_type": leave_type}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"刪除請假記錄失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# CodeTable (參數檔) API 端點
# ============================================

@app.get("/api/codes")
def get_code_tables(
    code_code: Optional[str] = Query(None, description="主分類代碼篩選"),
    used_mark: Optional[str] = Query(None, description="使用標記篩選"),
    page: int = Query(1, ge=1, description="頁碼"),
    page_size: int = Query(20, ge=1, le=100, description="每頁筆數"),
    sort_by: Optional[str] = Query(None, description="排序欄位"),
    sort_order: Optional[str] = Query("asc", description="排序方向 asc/desc"),
):
    """
    取得參數檔清單
    支援分頁、排序和篩選
    """
    try:
        with get_cursor() as cursor:
            base_sql = "SELECT * FROM gen001_allcode WHERE 1=1"
            params = []

            if code_code:
                base_sql += " AND code_code = %s"
                params.append(code_code)

            if used_mark:
                base_sql += " AND used_mark = %s"
                params.append(used_mark)

            # 計算總筆數
            count_sql = f"SELECT COUNT(*) as total FROM ({base_sql}) as subquery"
            cursor.execute(count_sql, params)
            total = cursor.fetchone()['total']

            # 排序
            allowed_sort_fields = ['code_code', 'code_subcode', 'code_subname', 'upddate', 'used_mark']
            if sort_by and sort_by in allowed_sort_fields:
                order_direction = 'ASC' if sort_order == 'asc' else 'DESC'
                base_sql += f" ORDER BY {sort_by} {order_direction}"
            else:
                base_sql += " ORDER BY code_code, code_subcode"

            # 分頁
            offset = (page - 1) * page_size
            base_sql += " LIMIT %s OFFSET %s"
            params.extend([page_size, offset])

            cursor.execute(base_sql, params)
            rows = cursor.fetchall()

            total_pages = (total + page_size - 1) // page_size

            return {
                "items": [dict(row) for row in rows],
                "total": total,
                "page": page,
                "page_size": page_size,
                "total_pages": total_pages,
            }

    except Exception as e:
        logger.error(f"取得參數檔失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/codes/leave-types", response_model=List[CodeTable])
def get_leave_types():
    """
    取得假別清單（code_code='0001'）
    """
    try:
        with get_cursor() as cursor:
            cursor.execute("""
                SELECT * FROM gen001_allcode 
                WHERE code_code = '0001' AND used_mark = '1'
                ORDER BY code_subcode
            """)
            rows = cursor.fetchall()
            return [dict(row) for row in rows]

    except Exception as e:
        logger.error(f"取得假別清單失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/codes/{code_code}/{code_subcode}", response_model=CodeTable)
def get_code_table(code_code: str, code_subcode: str):
    """
    根據複合主鍵取得單一參數
    """
    try:
        with get_cursor() as cursor:
            cursor.execute(
                "SELECT * FROM gen001_allcode WHERE code_code = %s AND code_subcode = %s",
                (code_code, code_subcode)
            )
            row = cursor.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="參數不存在")
            return dict(row)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"取得參數失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/codes", response_model=CodeTable)
def create_code_table(code: CodeTableCreate):
    """
    新增參數
    """
    from datetime import datetime
    try:
        with get_cursor() as cursor:
            # 檢查是否已存在
            cursor.execute(
                "SELECT code_code FROM gen001_allcode WHERE code_code = %s AND code_subcode = %s",
                (code.code_code, code.code_subcode)
            )
            if cursor.fetchone():
                raise HTTPException(status_code=400, detail="此參數已存在")

            # 取得非 None 的欄位
            data = code.model_dump(exclude_none=True)
            
            # 自動設定更新日期和時間
            now = datetime.now()
            data['upddate'] = now.strftime('%Y%m%d')
            data['updtime'] = now.strftime('%H%M%S')
            
            columns = ", ".join(data.keys())
            placeholders = ", ".join(["%s"] * len(data))
            values = list(data.values())

            sql = f"INSERT INTO gen001_allcode ({columns}) VALUES ({placeholders}) RETURNING *"
            cursor.execute(sql, values)
            row = cursor.fetchone()
            return dict(row)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"新增參數失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/codes/{code_code}/{code_subcode}", response_model=CodeTable)
def update_code_table(code_code: str, code_subcode: str, code: CodeTableUpdate):
    """
    更新參數
    """
    from datetime import datetime
    try:
        with get_cursor() as cursor:
            # 檢查是否存在
            cursor.execute(
                "SELECT code_code FROM gen001_allcode WHERE code_code = %s AND code_subcode = %s",
                (code_code, code_subcode)
            )
            if not cursor.fetchone():
                raise HTTPException(status_code=404, detail="參數不存在")

            # 取得非 None 的欄位進行更新
            data = code.model_dump(exclude_none=True)
            
            # 自動設定更新日期和時間
            now = datetime.now()
            data['upddate'] = now.strftime('%Y%m%d')
            data['updtime'] = now.strftime('%H%M%S')

            set_clause = ", ".join([f"{k} = %s" for k in data.keys()])
            values = list(data.values()) + [code_code, code_subcode]

            sql = f"UPDATE gen001_allcode SET {set_clause} WHERE code_code = %s AND code_subcode = %s RETURNING *"
            cursor.execute(sql, values)
            row = cursor.fetchone()
            return dict(row)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新參數失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/codes/{code_code}/{code_subcode}")
def delete_code_table(code_code: str, code_subcode: str):
    """
    刪除參數
    """
    try:
        with get_cursor() as cursor:
            cursor.execute(
                "DELETE FROM gen001_allcode WHERE code_code = %s AND code_subcode = %s RETURNING code_code",
                (code_code, code_subcode)
            )
            row = cursor.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="參數不存在")
            return {"message": "刪除成功", "code_code": code_code, "code_subcode": code_subcode}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"刪除參數失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================
# AnnualLeave (年度休假) API 端點
# ============================================

@app.get("/api/annual-leave")
def get_annual_leave_records(
    emp_id: Optional[str] = Query(None, description="員工編號篩選"),
    emp_name: Optional[str] = Query(None, description="員工姓名篩選"),
    year: Optional[str] = Query(None, description="年度篩選"),
    leave_type: Optional[str] = Query(None, description="假別篩選"),
    page: int = Query(1, ge=1, description="頁碼"),
    page_size: int = Query(20, ge=1, le=100, description="每頁筆數"),
    sort_by: Optional[str] = Query(None, description="排序欄位"),
    sort_order: Optional[str] = Query("desc", description="排序方向 asc/desc"),
):
    """
    取得年度休假記錄清單
    支援分頁、排序和篩選
    """
    try:
        with get_cursor() as cursor:
            # 基礎查詢（JOIN 員工資料取得姓名）
            base_sql = """
                SELECT a.*, m.chinese_name, m.name as english_name
                FROM member_annual_leave a
                LEFT JOIN member m ON a.emp_id = m.emp_id
                WHERE 1=1
            """
            params = []

            if emp_id:
                base_sql += " AND a.emp_id = %s"
                params.append(emp_id)

            if emp_name:
                base_sql += " AND (m.chinese_name LIKE %s OR m.name LIKE %s)"
                params.append(f"%{emp_name}%")
                params.append(f"%{emp_name}%")

            if year:
                base_sql += " AND a.year = %s"
                params.append(year)

            if leave_type:
                base_sql += " AND a.leave_type = %s"
                params.append(leave_type)

            # 計算總筆數
            count_sql = f"SELECT COUNT(*) as total FROM ({base_sql}) as subquery"
            cursor.execute(count_sql, params)
            total = cursor.fetchone()['total']

            # 排序
            allowed_sort_fields = ['emp_id', 'year', 'leave_type', 'days_of_leave', 'chinese_name']
            if sort_by and sort_by in allowed_sort_fields:
                order_direction = 'ASC' if sort_order == 'asc' else 'DESC'
                base_sql += f" ORDER BY {sort_by} {order_direction}"
            else:
                base_sql += " ORDER BY a.year DESC, a.emp_id"

            # 分頁
            offset = (page - 1) * page_size
            base_sql += " LIMIT %s OFFSET %s"
            params.extend([page_size, offset])

            cursor.execute(base_sql, params)
            rows = cursor.fetchall()

            total_pages = (total + page_size - 1) // page_size

            return {
                "items": [dict(row) for row in rows],
                "total": total,
                "page": page,
                "page_size": page_size,
                "total_pages": total_pages,
            }

    except Exception as e:
        logger.error(f"取得年度休假記錄失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/annual-leave/{emp_id}/{year}/{leave_type}")
def get_annual_leave_record(emp_id: str, year: str, leave_type: str):
    """
    根據複合主鍵取得單一年度休假記錄
    """
    try:
        with get_cursor() as cursor:
            cursor.execute(
                """
                SELECT a.*, m.chinese_name, m.name as english_name
                FROM member_annual_leave a
                LEFT JOIN member m ON a.emp_id = m.emp_id
                WHERE a.emp_id = %s AND a.year = %s AND a.leave_type = %s
                """,
                (emp_id, year, leave_type)
            )
            row = cursor.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="年度休假記錄不存在")
            return dict(row)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"取得年度休假記錄失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/annual-leave", status_code=201)
def create_annual_leave_record(record: AnnualLeaveCreate):
    """
    新增年度休假記錄
    """
    try:
        # 驗證可休天數：小數位只能是 0 或 0.5
        if record.days_of_leave is not None:
            decimal_part = record.days_of_leave % 1
            if decimal_part != 0 and decimal_part != 0.5:
                raise HTTPException(
                    status_code=400,
                    detail="可休天數的小數位僅能為 0 或 0.5"
                )

        with get_cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO member_annual_leave (emp_id, year, leave_type, days_of_leave, remark)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING emp_id, year, leave_type
                """,
                (
                    record.emp_id,
                    record.year,
                    record.leave_type,
                    record.days_of_leave,
                    record.remark,
                )
            )
            result = cursor.fetchone()
            return {
                "message": "新增成功",
                "emp_id": result['emp_id'],
                "year": result['year'],
                "leave_type": result['leave_type']
            }

    except HTTPException:
        raise
    except Exception as e:
        if "duplicate key" in str(e).lower():
            raise HTTPException(status_code=409, detail="該員工的此年度假別記錄已存在")
        logger.error(f"新增年度休假記錄失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/annual-leave/{emp_id}/{year}/{leave_type}")
def update_annual_leave_record(emp_id: str, year: str, leave_type: str, record: AnnualLeaveUpdate):
    """
    更新年度休假記錄
    """
    try:
        # 驗證可休天數：小數位只能是 0 或 0.5
        if record.days_of_leave is not None:
            decimal_part = record.days_of_leave % 1
            if decimal_part != 0 and decimal_part != 0.5:
                raise HTTPException(
                    status_code=400,
                    detail="可休天數的小數位僅能為 0 或 0.5"
                )

        update_data = record.model_dump(exclude_unset=True)
        if not update_data:
            raise HTTPException(status_code=400, detail="無更新欄位")

        with get_cursor() as cursor:
            set_clause = ", ".join([f"{k} = %s" for k in update_data.keys()])
            values = list(update_data.values())
            values.extend([emp_id, year, leave_type])

            cursor.execute(
                f"""
                UPDATE member_annual_leave
                SET {set_clause}
                WHERE emp_id = %s AND year = %s AND leave_type = %s
                RETURNING emp_id
                """,
                values
            )
            row = cursor.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="年度休假記錄不存在")
            return {"message": "更新成功", "emp_id": emp_id, "year": year, "leave_type": leave_type}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新年度休假記錄失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/annual-leave/{emp_id}/{year}/{leave_type}")
def delete_annual_leave_record(emp_id: str, year: str, leave_type: str):
    """
    刪除年度休假記錄
    """
    try:
        with get_cursor() as cursor:
            cursor.execute(
                "DELETE FROM member_annual_leave WHERE emp_id = %s AND year = %s AND leave_type = %s RETURNING emp_id",
                (emp_id, year, leave_type)
            )
            row = cursor.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="年度休假記錄不存在")
            return {"message": "刪除成功", "emp_id": emp_id, "year": year, "leave_type": leave_type}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"刪除年度休假記錄失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================
# ProjectInfo (專案資訊) API
# ============================================

class ImportMode(str, Enum):
    """匯入模式"""
    DELETE_ALL = "delete_all"  # 全部刪除後新增
    INSERT_ONLY = "insert_only"  # 僅新增不存在
    UPSERT = "upsert"  # 存在更新，不存在新增


# Excel 欄位名稱對應資料庫欄位
# NOTE: 欄位名稱與資料表 project_info 一致
EXCEL_COLUMN_MAPPING = {
    "專案代號": "project_id",
    "合約代號": "so_no",
    "專案名稱": "project_name",
    "客戶名稱": "customer_name",
    "專案計畫開始日": "project_plan_sdate",
    "專案計畫結束日": "project_plan_edate",
    "約定驗收日": "agreed_acceptance_date",
    "預計驗收日": "estimated_acceptance_date",
    "實際驗收日": "actual_acceptance_date",
    "保固開始日": "warranty_sdate",
    "保固結日日": "warranty_edate",
    "保固結束日": "warranty_edate",  # NOTE: 兩種可能的寫法
    "專案金額": "project_amt",
    "專案歸屬部門": "project_department",
    "專案負責人": "project_manager",
    "專案狀態": "project_status",
    "專案實際進度": "actual_progress",
    "專案收入": "project_income",
    "專案實際成本": "actual_cost",
    "專案類別": "project_category",
    "專案計畫進度": "project_plan_progress",
    "進度": "progress_status",
    "人力": "manpower_status",
    "品質": "quality_status",
    "計畫": "plan_status",
    "是否罰則": "is_penalty",
    "開發/維護階段預估人月": "estimated_dev_person_month",
    "全案實際人月": "actual_person_month",
    "預估保固成本": "estimated_warranty_cost",
    "預估保固人月": "estimated_warranty_person_month",
    "保固階段實際成本": "actual_warranty_cost",
    "保固階段實際人月": "actual_warranty_person_month",
}


@app.get("/api/projects")
def get_projects(
    project_id: Optional[str] = Query(None, description="專案代號篩選"),
    project_name: Optional[str] = Query(None, description="專案名稱篩選"),
    customer_name: Optional[str] = Query(None, description="客戶名稱篩選"),
    project_status: Optional[str] = Query(None, description="專案狀態篩選"),
    project_manager: Optional[str] = Query(None, description="專案負責人篩選"),
    date_from: Optional[str] = Query(None, description="專案開始日期（起）"),
    date_to: Optional[str] = Query(None, description="專案結束日期（迄）"),
    page: int = Query(1, ge=1, description="頁碼"),
    page_size: int = Query(20, ge=1, le=100, description="每頁筆數"),
    sort_by: Optional[str] = Query(None, description="排序欄位"),
    sort_order: Optional[str] = Query("desc", description="排序方向 asc/desc"),
):
    """
    取得專案列表
    """
    try:
        with get_cursor() as cursor:
            # 構建 WHERE 子句
            conditions = []
            params = []

            if project_id:
                conditions.append("project_id LIKE %s")
                params.append(f"%{project_id}%")
            if project_name:
                conditions.append("project_name LIKE %s")
                params.append(f"%{project_name}%")
            if customer_name:
                conditions.append("customer_name LIKE %s")
                params.append(f"%{customer_name}%")
            if project_status:
                # NOTE: project_status 為 char 類型，需使用 TRIM 去除尾部空格
                conditions.append("TRIM(project_status) = %s")
                params.append(project_status)
            if project_manager:
                conditions.append("project_manager LIKE %s")
                params.append(f"%{project_manager}%")
            # 日期篩選：專案計畫開始日 >= date_from
            if date_from:
                conditions.append("project_plan_sdate >= %s")
                params.append(date_from)
            # 日期篩選：專案計畫結束日 <= date_to
            if date_to:
                conditions.append("project_plan_edate <= %s")
                params.append(date_to)

            where_clause = " AND ".join(conditions) if conditions else "1=1"


            # 計算總數
            count_sql = f"SELECT COUNT(*) as cnt FROM project_info WHERE {where_clause}"
            cursor.execute(count_sql, params)
            total = cursor.fetchone()["cnt"]

            # 排序
            valid_sort_columns = [
                "project_id", "project_name", "customer_name", "project_status",
                "project_manager", "project_amt", "project_plan_start", "project_plan_end",
                "actual_progress", "created_at", "updated_at"
            ]
            order_clause = "project_id ASC"
            if sort_by and sort_by in valid_sort_columns:
                direction = "DESC" if sort_order and sort_order.lower() == "desc" else "ASC"
                order_clause = f"{sort_by} {direction}"

            # 分頁
            offset = (page - 1) * page_size
            query = f"""
                SELECT * FROM project_info
                WHERE {where_clause}
                ORDER BY {order_clause}
                LIMIT %s OFFSET %s
            """
            cursor.execute(query, params + [page_size, offset])
            rows = cursor.fetchall()

            items = []
            for row in rows:
                item = dict(row)
                # 處理 Decimal 和日期類型
                for key, value in item.items():
                    if hasattr(value, 'isoformat'):
                        item[key] = value.isoformat()
                    elif hasattr(value, '__float__'):
                        item[key] = float(value)
                items.append(item)

            return {
                "items": items,
                "total": total,
                "page": page,
                "page_size": page_size,
                "total_pages": (total + page_size - 1) // page_size,
            }

    except Exception as e:
        logger.error(f"查詢專案列表失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/projects/filter-options")
def get_project_filter_options():
    """
    取得專案過濾器選項
    """
    try:
        with get_cursor() as cursor:
            # 取得所有狀態
            cursor.execute("SELECT DISTINCT project_status FROM project_info WHERE project_status IS NOT NULL ORDER BY project_status")
            statuses = [row["project_status"] for row in cursor.fetchall()]
            
            # 取得所有專案
            cursor.execute("SELECT project_id, project_name FROM project_info ORDER BY project_name")
            projects = [{"id": row["project_id"], "name": row["project_name"]} for row in cursor.fetchall()]
            
            return {"statuses": statuses, "projects": projects}
    except Exception as e:
        logger.error(f"取得過濾選項失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/projects/stats")
def get_project_stats(
    dimension: str = Query("status", description="統計維度: status, customer, department"),
    interval: str = Query("none", description="統計區間: none, monthly, quarterly, yearly"),
    date_from: Optional[str] = Query(None, description="起始日期 (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="結束日期 (YYYY-MM-DD)"),
    statuses: Optional[str] = Query(None, description="專案狀態（逗號分隔）"),
):
    """
    取得專案統計資料
    NOTE: 支援日期範圍、狀態多選、時間區間分組
    """
    try:
        with get_cursor() as cursor:
            # 建立篩選條件
            conditions = []
            params = []
            
            # NOTE: project_plan_sdate 是字串類型 (YYYY-MM-DD)
            
            # 日期範圍篩選
            if date_from:
                conditions.append("project_plan_sdate >= %s")
                params.append(date_from)
            if date_to:
                conditions.append("project_plan_sdate <= %s")
                params.append(date_to)
            
            # 專案狀態多選
            if statuses:
                status_list = [s.strip() for s in statuses.split(",") if s.strip()]
                if status_list:
                    placeholders = ", ".join(["%s"] * len(status_list))
                    conditions.append(f"project_status IN ({placeholders})")
                    params.extend(status_list)
            
            where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""
            
            # 根據區間選擇時間分組欄位
            if interval == "monthly":
                time_group = "SUBSTRING(project_plan_sdate, 1, 7)"  # YYYY-MM
            elif interval == "quarterly":
                time_group = "SUBSTRING(project_plan_sdate, 1, 4) || '-Q' || ((CAST(SUBSTRING(project_plan_sdate, 6, 2) AS INT) - 1) / 3 + 1)::TEXT"
            elif interval == "yearly":
                time_group = "SUBSTRING(project_plan_sdate, 1, 4)"  # YYYY
            else:
                time_group = None
            
            # 根據維度選擇分組欄位
            if dimension == "status":
                dim_group = "COALESCE(project_status, '未設定')"
            elif dimension == "customer":
                dim_group = "COALESCE(customer_name, '未設定')"
            elif dimension == "department":
                dim_group = "COALESCE(project_department, '未設定')"
            else:
                dim_group = "COALESCE(project_status, '未設定')"
            
            # 組合分組欄位
            if time_group:
                # 有時間區間：時間 + 維度
                group_field = f"{time_group} || ' - ' || {dim_group}"
                order_field = "name"
            else:
                # 無時間區間：只用維度
                group_field = dim_group
                order_field = "count DESC"
            
            query = f"""
                SELECT 
                    {group_field} as name,
                    COUNT(*) as count,
                    COALESCE(SUM(project_amt), 0) as amount
                FROM project_info
                {where_clause}
                GROUP BY {group_field}
                ORDER BY {order_field}
            """
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            
            items = [
                {
                    "name": row["name"] if row["name"] else "未設定",
                    "count": row["count"],
                    "amount": float(row["amount"]) if row["amount"] else 0
                }
                for row in rows
            ]
            
            return {"items": items}
            
    except Exception as e:
        logger.error(f"統計專案失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/projects/{project_id}")
def get_project(project_id: str):
    """
    取得單一專案
    """
    try:
        with get_cursor() as cursor:
            cursor.execute("SELECT * FROM project_info WHERE project_id = %s", (project_id,))
            row = cursor.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="專案不存在")
            item = dict(row)
            for key, value in item.items():
                if hasattr(value, 'isoformat'):
                    item[key] = value.isoformat()
                elif hasattr(value, '__float__'):
                    item[key] = float(value)
            return item

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"查詢專案失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/projects", status_code=201)
def create_project(project: ProjectInfoCreate):
    """
    新增專案
    """
    try:
        with get_cursor() as cursor:
            # 檢查是否已存在
            cursor.execute("SELECT 1 FROM project_info WHERE project_id = %s", (project.project_id,))
            if cursor.fetchone():
                raise HTTPException(status_code=400, detail="專案代號已存在")

            # 構建欄位和值
            data = project.model_dump(exclude_none=True)
            columns = list(data.keys())
            placeholders = ", ".join(["%s"] * len(columns))
            values = [data[col] for col in columns]

            sql = f"""
                INSERT INTO project_info ({', '.join(columns)})
                VALUES ({placeholders})
            """
            cursor.execute(sql, values)
            return {"message": "新增成功", "project_id": project.project_id}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"新增專案失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/projects/{project_id}")
def update_project(project_id: str, project: ProjectInfoUpdate):
    """
    更新專案
    """
    try:
        with get_cursor() as cursor:
            # 檢查是否存在
            cursor.execute("SELECT 1 FROM project_info WHERE project_id = %s", (project_id,))
            if not cursor.fetchone():
                raise HTTPException(status_code=404, detail="專案不存在")

            # 構建更新語句
            data = project.model_dump(exclude_none=True)
            if not data:
                raise HTTPException(status_code=400, detail="沒有要更新的欄位")

            set_clause = ", ".join([f"{k} = %s" for k in data.keys()])
            values = list(data.values()) + [project_id]

            sql = f"UPDATE project_info SET {set_clause}, updated_at = NOW() WHERE project_id = %s"
            cursor.execute(sql, values)
            return {"message": "更新成功"}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新專案失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/projects/{project_id}")
def delete_project(project_id: str):
    """
    刪除專案
    """
    try:
        with get_cursor() as cursor:
            cursor.execute(
                "DELETE FROM project_info WHERE project_id = %s RETURNING project_id",
                (project_id,)
            )
            if not cursor.fetchone():
                raise HTTPException(status_code=404, detail="專案不存在")
            return {"message": "刪除成功", "project_id": project_id}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"刪除專案失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/projects/import")
async def import_projects(
    file: UploadFile = File(...),
    mode: str = Form(...),
):
    """
    匯入專案資料
    支援 Excel 格式 (.xlsx, .xls)

    mode:
    - delete_all: 全部刪除後新增
    - insert_only: 僅新增不存在
    - upsert: 存在更新，不存在新增
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="需要安裝 openpyxl 模組")

    # 驗證檔案類型
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="只支援 Excel 檔案 (.xlsx, .xls)")

    # 驗證匯入模式
    if mode not in ["delete_all", "insert_only", "upsert"]:
        raise HTTPException(status_code=400, detail="無效的匯入模式")

    try:
        # 讀取 Excel 檔案
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        sheet = workbook.active

        # 找到標題行（搜尋「專案代號」所在位置）
        header_row = None
        header_col_start = None
        column_mapping = {}

        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20), start=1):
            for col_idx, cell in enumerate(row, start=1):
                cell_value = str(cell.value).strip() if cell.value else ""
                if cell_value == "專案代號":
                    header_row = row_idx
                    header_col_start = col_idx
                    break
            if header_row:
                break

        if not header_row:
            raise HTTPException(status_code=400, detail="找不到標題行（需包含「專案代號」欄位）")

        # 建立欄位對應
        for col_idx, cell in enumerate(
            sheet.iter_rows(min_row=header_row, max_row=header_row, min_col=header_col_start).__next__(),
            start=header_col_start
        ):
            cell_value = str(cell.value).strip() if cell.value else ""
            if cell_value in EXCEL_COLUMN_MAPPING:
                column_mapping[col_idx] = EXCEL_COLUMN_MAPPING[cell_value]

        if "project_id" not in column_mapping.values():
            raise HTTPException(status_code=400, detail="找不到專案代號欄位")

        # 解析資料行
        records = []
        for row in sheet.iter_rows(min_row=header_row + 1, min_col=header_col_start):
            record = {}
            has_data = False
            for col_idx, cell in enumerate(row, start=header_col_start):
                if col_idx in column_mapping:
                    field_name = column_mapping[col_idx]
                    value = cell.value
                    if value is not None:
                        has_data = True
                        # 處理不同類型的值
                        if field_name == "is_penalty":
                            record[field_name] = bool(value) if value else False
                        elif field_name in ["project_amt", "actual_cost", "actual_progress",
                                            "project_plan_progress", "development_person",
                                            "estimated_dev_person", "actual_person_month",
                                            "estimated_warranty", "estimated_warranty_person",
                                            "actual_warranty", "actual_warranty_person"]:
                            try:
                                # 處理百分比格式
                                if isinstance(value, str) and "%" in value:
                                    record[field_name] = float(value.replace("%", "").strip())
                                else:
                                    record[field_name] = float(value) if value else None
                            except (ValueError, TypeError):
                                record[field_name] = None
                        elif hasattr(value, 'strftime'):
                            # 日期格式
                            record[field_name] = value.strftime("%Y/%m/%d")
                        else:
                            record[field_name] = str(value).strip() if value else None

            if has_data and record.get("project_id"):
                records.append(record)

        if not records:
            raise HTTPException(status_code=400, detail="沒有找到有效的資料")

        # 執行匯入
        with get_cursor() as cursor:
            inserted = 0
            updated = 0
            skipped = 0

            if mode == "delete_all":
                # 全部刪除後新增
                cursor.execute("DELETE FROM project_info")
                for record in records:
                    columns = list(record.keys())
                    placeholders = ", ".join(["%s"] * len(columns))
                    values = [record[col] for col in columns]
                    sql = f"INSERT INTO project_info ({', '.join(columns)}) VALUES ({placeholders})"
                    cursor.execute(sql, values)
                    inserted += 1

            elif mode == "insert_only":
                # 僅新增不存在的
                for record in records:
                    project_id = record.get("project_id")
                    cursor.execute("SELECT 1 FROM project_info WHERE project_id = %s", (project_id,))
                    if cursor.fetchone():
                        skipped += 1
                        continue
                    columns = list(record.keys())
                    placeholders = ", ".join(["%s"] * len(columns))
                    values = [record[col] for col in columns]
                    sql = f"INSERT INTO project_info ({', '.join(columns)}) VALUES ({placeholders})"
                    cursor.execute(sql, values)
                    inserted += 1

            elif mode == "upsert":
                # 存在更新，不存在新增
                for record in records:
                    project_id = record.get("project_id")
                    cursor.execute("SELECT 1 FROM project_info WHERE project_id = %s", (project_id,))
                    if cursor.fetchone():
                        # 更新
                        update_data = {k: v for k, v in record.items() if k != "project_id"}
                        if update_data:
                            set_clause = ", ".join([f"{k} = %s" for k in update_data.keys()])
                            values = list(update_data.values()) + [project_id]
                            sql = f"UPDATE project_info SET {set_clause}, updated_at = NOW() WHERE project_id = %s"
                            cursor.execute(sql, values)
                            updated += 1
                    else:
                        # 新增
                        columns = list(record.keys())
                        placeholders = ", ".join(["%s"] * len(columns))
                        values = [record[col] for col in columns]
                        sql = f"INSERT INTO project_info ({', '.join(columns)}) VALUES ({placeholders})"
                        cursor.execute(sql, values)
                        inserted += 1

        return {
            "message": "匯入完成",
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
            "total": len(records),
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"匯入專案資料失敗: {e}")
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
